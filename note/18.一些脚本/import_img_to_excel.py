import os
import glob
from openpyxl import Workbook
from openpyxl.drawing.image import Image as OpenpyxlImage
try:
    from PIL import Image as PILImage
    PIL_AVAILABLE = True
except ImportError:
    PIL_AVAILABLE = False
    print("警告: PIL/Pillow未安装，将使用原始图片大小")

def resize_image_if_needed(img_path, max_width=200, max_height=150):
    """调整图片大小，如果PIL可用的话"""
    if not PIL_AVAILABLE:
        return img_path, None
    
    try:
        with PILImage.open(img_path) as pil_img:
            width, height = pil_img.size
            
            # 如果图片已经足够小，直接返回原路径
            if width <= max_width and height <= max_height:
                return img_path, (width, height)
            
            # 计算缩放比例
            scale = min(max_width/width, max_height/height)
            new_width = int(width * scale)
            new_height = int(height * scale)
            
            # 调整图片大小
            resized_img = pil_img.resize((new_width, new_height), PILImage.Resampling.LANCZOS)
            
            # 保存临时调整后的图片
            temp_path = f"temp_resized_{os.path.basename(img_path)}"
            resized_img.save(temp_path, quality=95)
            
            return temp_path, (new_width, new_height)
    except Exception as e:
        print(f"调整图片大小时出错: {e}")
        return img_path, None

def is_image_file(filename):
    """检查文件是否为图片格式"""
    image_extensions = {
        '.jpg', '.jpeg', '.png', '.gif', '.bmp', '.tiff', '.tif',
        '.webp', '.ico', '.svg', '.psd', '.raw', '.heic', '.heif'
    }
    
    # 获取文件扩展名并转为小写
    _, ext = os.path.splitext(filename.lower())
    return ext in image_extensions

def create_image_excel():
    """
    创建包含当前目录下所有图片的Excel文件
    图片名称在A列，对应图片在B列
    只处理图片格式文件，忽略其他所有格式
    """
    
    # 获取当前目录下的所有文件
    all_files = [f for f in os.listdir('.') if os.path.isfile(f)]
    
    # 筛选出图片文件
    image_files = [f for f in all_files if is_image_file(f)]
    
    if not image_files:
        print("当前目录下没有找到图片文件")
        return
    
    print(f"找到 {len(image_files)} 个图片文件")
    
    # 创建工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "图片目录"
    
    # 设置列标题
    ws['A1'] = "图片名称"
    ws['B1'] = "图片"
    
    # 设置列宽和标题行样式
    ws.column_dimensions['A'].width = 35
    ws.column_dimensions['B'].width = 25
    ws.row_dimensions[1].height = 20
    
    # 记录需要清理的临时文件
    temp_files = []
    
    try:
        # 遍历图片文件
        for idx, img_path in enumerate(sorted(image_files), start=2):
            try:
                # 检查文件是否存在
                if not os.path.exists(img_path):
                    print(f"文件不存在: {img_path}")
                    continue
                
                # 获取图片名称（不含扩展名）
                img_name = os.path.splitext(os.path.basename(img_path))[0]
                
                # 在A列写入图片名称
                ws[f'A{idx}'] = img_name
                
                # 调整图片大小
                processed_path, dimensions = resize_image_if_needed(img_path)
                
                # 如果创建了临时文件，记录下来
                if processed_path != img_path:
                    temp_files.append(processed_path)
                
                # 创建openpyxl图片对象
                img_obj = OpenpyxlImage(processed_path)
                
                # 设置图片大小（如果有尺寸信息）
                if dimensions:
                    img_obj.width = dimensions[0]
                    img_obj.height = dimensions[1]
                else:
                    # 默认大小限制
                    if hasattr(img_obj, 'width') and img_obj.width > 200:
                        scale = 200 / img_obj.width
                        img_obj.width = 200
                        img_obj.height = int(img_obj.height * scale)
                    if hasattr(img_obj, 'height') and img_obj.height > 150:
                        scale = 150 / img_obj.height
                        img_obj.height = 150
                        img_obj.width = int(img_obj.width * scale)
                
                # 设置图片位置（B列对应的单元格）
                img_obj.anchor = f'B{idx}'
                
                # 将图片添加到工作表
                ws.add_image(img_obj)
                
                # 设置行高以适应图片
                if dimensions:
                    row_height = max(dimensions[1] * 0.75, 25)
                else:
                    row_height = 100
                ws.row_dimensions[idx].height = min(row_height, 150)
                
                print(f"已处理: {img_name}")
                
            except Exception as e:
                print(f"处理图片 {img_path} 时出错: {str(e)}")
                continue
        
        # 保存Excel文件
        output_file = "图片目录.xlsx"
        wb.save(output_file)
        print(f"\nExcel文件已保存为: {output_file}")
        print(f"共处理了 {len([f for f in image_files if os.path.exists(f)])} 个图片文件")
        
    finally:
        # 清理临时文件
        for temp_file in temp_files:
            try:
                if os.path.exists(temp_file):
                    os.remove(temp_file)
            except Exception as e:
                print(f"删除临时文件 {temp_file} 时出错: {e}")

if __name__ == "__main__":
    create_image_excel()